# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import ValidationError
from datetime import datetime
import os
import base64
import io

directory = os.path.dirname(__file__)
# directory = os.path.dirname("/home/odoo/data/filestore/xls")  # For Odoo SH

try:
    from openpyxl import load_workbook
except ImportError:
    pass


class TvaDeclaration(models.Model):
    _name = "tva.declaration"
    _description = "Declaration Tva"
    _inherit = ['mail.thread']

    @api.constrains('period_id')
    def _check_period_unicity(self):
        tva_ids = self.search([('period_id', '=', self.period_id.id), ('id', '!=', self.id)])
        if tva_ids:
            raise ValidationError(u"Une autre déclaration TVA pour la même période existe déjà!")

    @api.constrains('state')
    def _check_closed_tva(self):
        tva_ids = self.search([('state', '=', 'draft'), ('id', '!=', self.id)])
        if tva_ids:
            raise ValidationError(u"Toutes les autres déclarations de TVA doivent être cloturées!")

    @api.depends('period_id')
    def get_name(self):
        for record in self:
            name = u'Déclaration de TVA: '
            if self.period_id:
                name += record.period_id.name
            record.name = name

    name = fields.Char('Description', compute='get_name', store=True)
    period_id = fields.Many2one('date.range',
                                string=u'Période', domain=[('type_id.fiscal_period', '=', True)], required=True)
    company_id = fields.Many2one('res.company', readonly=True, string=u'Societé',
                                default=lambda self: self.env['res.company']._company_default_get('tva.declaration'))
    line_ids = fields.One2many(comodel_name='tva.declaration.line', inverse_name='tva_declaration_id',
                               string=u'Décaissements', domain=[('type', '=', False)])
    encaissement_ids = fields.One2many(comodel_name='tva.declaration.line', inverse_name='tva_declaration_id',
                                       string=u'Encaissements', domain=[('type', '=', True)])

    receivable_tva_ids = fields.One2many(comodel_name='tva.line', inverse_name='tva_declaration_id',
                                         string=u'TVA fournisseur', domain=[('type', '=', 'p')],
                                         compute="generate_tva_lines")
    payable_tva_ids = fields.One2many(comodel_name='tva.line', inverse_name='tva_declaration_id',
                                      string=u'TVA Client', domain=[('type', '=', 'r')], compute="generate_tva_lines")

    state = fields.Selection([
        ('draft', u'Brouillon'),
        ('done', u'Valide')], default='draft', string='Etat', readonly=True, track_visibility='onchange')

    tva_report_file = fields.Binary('Fichier Excel', attachment=True)
    name_file_excel = fields.Char(string=u"Nom fichier", required=False, )
    payment_proof_file = fields.Binary(u'Reçu de paiement', attachment=True)
    regime = fields.Char(u'Régime', compute="get_data_tva")
    annee = fields.Char(u'Année', compute="get_data_tva")
    mois = fields.Char(u'Mois', compute="get_data_tva", store=True)
    period = fields.Char(u'Période', compute="get_data_tva")
    tax_account_id = fields.Many2one('account.account', string=u"Compte Crédit de TVA",
                                     default=lambda self: self.env.user.company_id.tax_account_id, required=True)

    payed_tax_account_id = fields.Many2one('account.account', string=u"Compte TVA due",
                                           default=lambda self: self.env.user.company_id.payed_tax_account_id,
                                           required=True)

    # Champs Compute TVA à payer
    sum_receivable = fields.Float('Total TVA collectée', compute="get_tva_amount")
    sum_payable = fields.Float('Total TVA déductible', compute="get_tva_amount")
    to_be_payed = fields.Float(u'TVA à déclarer', compute="get_tva_amount")
    sum_credit = fields.Float(u'Crédit de la période précédente')
    move_id = fields.Many2one('account.move', u'Pièce comptable', readonly=True)

    date = fields.Date('Date', required=True)

    code_tva_202 = fields.Float(u'Réduction de 15% du crédit de la période ((190 - 170) - 130) × 15%')
    code_tva_203 = fields.Float(u'Crédit restant cumulé après réduction de 15% (201- 202)')
    code_tva_157 = fields.Float(
        u"Déduction prévue par l'article 125-VII du CGI (1/5 du montant de la TVA payée au cours du mois de décembre 2013)")

    manual_field_ids = fields.One2many('tva.compute.manual_line', 'tva_dec_id', string="Champs à saisie manuelle")
    auto_field_ids = fields.One2many('tva.compute.auto_line', 'tva_dec_id', string="Champs Calculés")

    credit_tva = fields.Float(u'Crédit de TVA', compute="get_tva_amount")
    cred_pay = fields.Float(u'Crédit accompagné de paiement (ligne 204)', compute="get_tva_amount")
    tva_due = fields.Float(u'TVA due', compute="get_tva_amount")

    def get_tva_credit(self):
        for record in self:
            account_mv_line_obj = self.env['account.move.line']
            tva_dec_obj = self.env['tva.declaration']
            sum_credit = 0
            moves = account_mv_line_obj.search([('account_id', '=', record.tax_account_id.id),
                                                ('full_reconcile_id', '=', False)
                                                ]).filtered(lambda r: r.move_id.state == 'posted')

            for mv in moves:
                sum_credit += mv.debit - mv.credit
            if sum_credit < 0:
                sum_credit = 0
            record.sum_credit = sum_credit
    @api.depends('line_ids','encaissement_ids')
    def get_tva_amount(self):
        for record in self:
            sum_receivable = 0
            sum_payable = 0
            record.credit_tva = 0
            record.tva_due = 0
            record.cred_pay = 0
            for line in record.line_ids:
                if not line.to_be_delayed:
                    sum_payable += line.amount_tva
            for line in record.encaissement_ids:
                if not line.to_be_delayed:
                    sum_receivable += line.amount_tva
            sum_etranger = sum(o.amount_tva for o in self.line_ids
                               .filtered(lambda r: r.tax_id.code in [145]))  # 129
            sum_201 = sum(o.value for o in self.auto_field_ids
                          .filtered(lambda r: r.tva_report_id.code in [201]))  # 201
            sum_200 = sum(o.value for o in self.auto_field_ids
                          .filtered(lambda r: r.tva_report_id.code in [200])) # 200
            sum_204 = sum(o.value for o in self.auto_field_ids
                          .filtered(lambda r: r.tva_report_id.code in [204]))  # 204
            record.sum_payable = sum_payable + sum_etranger
            record.sum_receivable = sum_receivable
            record.to_be_payed = sum_receivable - sum_payable - record.sum_credit
            record.credit_tva = sum_201
            record.cred_pay = sum_204
            record.tva_due = sum_200
            if sum_etranger > 0 and (sum_receivable - sum_etranger - sum_payable < 0):
                record.sum_payable = record.sum_payable - sum_etranger

            # if record.to_be_payed < 0:
            #     record.credit_tva = abs(record.to_be_payed)
            # else:
            #     record.tva_due = record.to_be_payed

    # Generate Accounting Move
    def genetare_tax_account_move(self):
        account_move_obj = self.env['account.move']
        for record in self:
            journal_id = self.env.user.company_id.tax_journal_id.id
            move_data = []
            record.move_id.unlink()
            # Décaissements
            for line in record.line_ids:
                if not line.to_be_delayed:
                    if line.amount_tva > 0.0:
                        move_data.append({
                            'name': record.name,
                            'date': record.date,
                            'journal_id': journal_id,
                            'account_id': line.tax_id.invoice_repartition_line_ids[1].account_id.id,
                            'credit': line.amount_tva,
                            # 'invoice_id': line.invoice_tax_id.invoice_id.id,
                            'move_id': line.move_id.id,
                        })
                    if line.amount_tva < 0.0:
                        move_data.append({
                            'name': record.name,
                            'date': record.date,
                            'journal_id': journal_id,
                            'account_id': line.tax_id.invoice_repartition_line_ids[1].account_id.id,
                            'debit': abs(line.amount_tva),
                            # 'invoice_id': line.invoice_tax_id.invoice_id.id,
                            'move_id': line.move_id.id,
                        })
            # Encaissements
            for line in record.encaissement_ids:
                if not line.to_be_delayed:
                    if line.amount_tva > 0.0:
                        move_data.append({
                            'name': record.name,
                            'journal_id': journal_id,
                            'date': record.date,
                            'account_id': line.tax_id.invoice_repartition_line_ids[1].account_id.id,
                            'debit': line.amount_tva,
                            # 'invoice_id': line.invoice_tax_id.invoice_id.id,
                            'move_id': line.move_id.id,
                        })
                    if line.amount_tva < 0.0:
                        move_data.append({
                            'name': record.name,
                            'journal_id': journal_id,
                            'date': record.date,
                            'account_id': line.tax_id.invoice_repartition_line_ids[1].account_id.id,
                            'credit': abs(line.amount_tva),
                            # 'invoice_id': line.invoice_tax_id.invoice_id.id,
                            'move_id': line.move_id.id,
                        })
            # En cas de crédit de TVA
            if record.to_be_payed < 0:
                move_data.append({
                    'name': record.name,
                    'date': record.date,
                    'journal_id': journal_id,
                    'account_id': record.tax_account_id.id,
                    'debit': abs(record.to_be_payed),
                })
                if record.sum_credit:
                    move_data.append({
                        'name': record.name,
                        'journal_id': journal_id,
                        'date': record.date,
                        'account_id': record.tax_account_id.id,
                        'credit': abs(record.sum_credit),
                    })
            # En cas de TVA due
            elif record.to_be_payed > 0:
                move_data.append({
                    'name': record.name,
                    'date': record.date,
                    'journal_id': journal_id,
                    'account_id': record.payed_tax_account_id.id,
                    'credit': record.to_be_payed,
                    # 'invoice_id': line.invoice_tax_id.invoice_id.id,
                    'move_id': line.move_id.id,
                })
                if record.sum_credit:
                    move_data.append({
                        'name': record.name,
                        'date': record.date,
                        'journal_id': journal_id,
                        'account_id': record.tax_account_id.id,
                        'credit': record.sum_credit,
                        # 'invoice_id': line.invoice_tax_id.invoice_id.id,
                        'move_id': line.move_id.id,
                    })

            line_ids = [(0, 0, mv) for mv in move_data]
            move_id = account_move_obj.create({
                'journal_id': journal_id,
                'date': record.date,
                'ref': record.name,
                'line_ids': line_ids,
            })
            record.move_id = move_id

    # Compute regime & period
    @api.depends('period_id')
    def get_data_tva(self):
        for record in self:
            if record.date:
                record.mois = record.date.month
            if record.period_id and record.period_id.fiscal_year_id:
                if len(record.period_id.fiscal_year_id.period_ids) == 12:
                    record.regime = "1"
                else:
                    record.regime = "2"
            if record.period_id:
                split_date = record.period_id.date_start
                record.annee = split_date.year
                if record.regime == "2":
                    record.period = str(((split_date.month - 1) // 3) + 1)
                else:
                    if split_date.month in ('10', '11', '12'):
                        record.period = split_date.month
                    else:
                        record.period = split_date.month
            else:
                record.regime = "1"
                record.annee = datetime.today().year
                record.period = str(((datetime.today().month - 1) // 3) + 1)

    def generate_tva_file(self):
        tax_report_cells = self.env['account.tax.repport'].search([])  # Lignes de rapport TVA
        tva_line_obj = self.env['tva.line']  # Lignes de TVA facturés dans la déclaration TVA
        tva_auto_line_obj = self.env['tva.compute.auto_line']
        tva_manu_line_obj = self.env['tva.compute.manual_line']
        report_template = self.env['ir.config_parameter'].sudo().get_param('tva_encaissement_maroc.tva_rapport_id')

        if not report_template:
            raise ValidationError(u'Merci de renseigner le fichier de la déclaration au niveau de la configuration')
        file = base64.b64decode(report_template)
        xls_filelike = io.BytesIO(file)
        wb = load_workbook(xls_filelike)
        # ws = wb.get_active_sheet()
        ws = wb.active
        cell_achat_import = self.env['account.tax.repport'].search([('code', '=', 129)])
        for record in self:
            # tax fourni etranger
            tax_external_supplier = tva_line_obj.search(
                [('tva_declaration_id', '=', record.id), ('tax_id.code', '=', 145)])
            for cell in tax_report_cells:
                line = tva_line_obj.search([('tva_declaration_id', '=', record.id), ('tax_id.code', '=', cell.code)])
                auto_line = tva_auto_line_obj.search(
                    [('tva_dec_id', '=', record.id), ('tva_report_id.code', '=', cell.code)])
                manu_line = tva_manu_line_obj.search(
                    [('tva_dec_id', '=', record.id), ('tva_report_id.code', '=', cell.code)])
                if cell.cell_base:
                    ws[cell.cell_base] = line.amount_ht or auto_line.value
                if cell.cell_tax:
                    ws[cell.cell_tax] = line.amount or manu_line.value

                if cell.cell_prorata:
                    prorata = int(line.tax_id.prorata) or 100
                    ws[cell.cell_prorata] = prorata / 100
                    if cell.cell_calc_tax:
                        amount_tva = line.amount or auto_line.value
                        ws[cell.cell_calc_tax] = amount_tva * (prorata / 100)

                # Cas du fournisseur étranger
                if cell.code == 129:
                    ws[cell_achat_import.cell_base] = tax_external_supplier.amount_ht
                    ws[cell_achat_import.cell_tax] = tax_external_supplier.amount

            wb.save(os.path.join(directory, 'tva.xlsx'))
            tva_report_file = base64.encodestring(open(os.path.join(directory, 'tva.xlsx'), 'rb').read())
            extension = 'xlsx'
            filename = u"Déclaration TVA-" + record.env.user.company_id.name + '-' + record.period + '-' + record.annee
            name = "%s.%s" % (filename, extension)
            record.write({'tva_report_file': tva_report_file, 'name_file_excel': name})

    def action_draft(self):
        for record in self:
            record.state = 'draft'

    def validate(self):
        for record in self:
            if not record.payment_proof_file:
                raise ValidationError(u"Merci d'attacher le reçu de paiement!")
            record.state = 'done'

    '''
        returns the payment type of every line in the invoice
    '''

    # Get Payment Type
    def get_payment(self, move):
        paiement_type = False
        for move_line in move.line_ids:
            if move_line.payement_method_id:
                paiement_type = move_line.payement_method_id.id
                break
        return paiement_type

    # Compute receivable_tva_ids & payable_tva_ids
    def generate_tva_lines(self):
        tva_obj = self.env['tva.line']
        receivable_tva_ids = tva_obj
        payable_tva_ids = tva_obj
        for record in self:
            tva_obj.search([('tva_declaration_id', '=', record.id)]).unlink()
            data = {}
            for line in record.line_ids:
                if not line.to_be_delayed:
                    if data.get(line.tax_id.id, False):
                        data[line.tax_id.id]['amount'] += line.amount_tva
                        data[line.tax_id.id]['amount_ht'] += line.amount_ht
                    else:
                        data[line.tax_id.id] = {
                            'tva_declaration_id': record.id,
                            'tax_id': line.tax_id.id,
                            'amount': line.amount_tva,
                            'amount_ht': line.amount_ht,
                            'type': 'r'
                        }
            for line in record.encaissement_ids:
                if not line.to_be_delayed:
                    if data.get(line.tax_id.id, False):
                        data[line.tax_id.id]['amount'] += line.amount_tva
                        data[line.tax_id.id]['amount_ht'] += line.amount_ht
                    else:
                        data[line.tax_id.id] = {
                            'tva_declaration_id': record.id,
                            'tax_id': line.tax_id.id,
                            'amount': line.amount_tva,
                            'amount_ht': line.amount_ht,
                            'type': 'p'
                        }
            for data_tva in data:
                if data[data_tva]['type'] == 'r':
                    receivable_tva_ids += tva_obj.create(data[data_tva])
                if data[data_tva]['type'] == 'p':
                    payable_tva_ids += tva_obj.create(data[data_tva])
            self.receivable_tva_ids = receivable_tva_ids
            self.payable_tva_ids = payable_tva_ids

    '''
        gets payments of partner/client invoices
    '''

    # Get Payment Counterpart
    def generate_liquidity_moves(self, move):
        move_pay_ids = []
        for move_line in move.line_ids:
            if move_line.debit > 0:
                reconcial_ids = self.env['account.partial.reconcile'].search(
                    [('debit_move_id', '=', move_line.id)])
                if reconcial_ids:
                    for rec in reconcial_ids:
                        move_pay_ids.append([rec.credit_move_id, rec.amount])
            if move_line.credit > 0:
                reconcial_ids = self.env['account.partial.reconcile'].search(
                    [('credit_move_id', '=', move_line.id)])
                if reconcial_ids:
                    for rec in reconcial_ids:
                        move_pay_ids.append([rec.debit_move_id, rec.amount])
        return move_pay_ids

    '''
        get moves in the period given 
    '''

    def get_payment_moves(self):
        move_obj = self.env['account.move']  # pièces
        domain = [('date', '>=', self.period_id.date_start),
                  ('date', '<=', self.period_id.date_end),
                  ('journal_id.type', 'in', ('cash', 'bank'))]
        move_ids = move_obj.search(domain)
        return move_ids

    '''
        Champs à calculer
    '''

    def get_auto_computed_lines(self):
        auto_lines = self.env['account.tax.repport'].search([('type_calcul', '=', 'auto')])
        self.auto_field_ids.unlink()
        for line in auto_lines:
            # on pourrait faire le calcul auto ici par ex
            sum_sale_ht = sum(l.amount_ht for l in self.encaissement_ids
                              .filtered(lambda r: r.to_be_delayed is False))  # Code 010
            tva_code_20 = sum(l.amount_ht for l in self.encaissement_ids
                              .filtered(lambda r: r.to_be_delayed is False and r.tax_id.code == 20))  # Code 20
            tva_code_30 = sum(l.amount_ht for l in self.encaissement_ids
                              .filtered(lambda r: r.to_be_delayed is False and r.tax_id.code == 30))  # Code 30
            tva_code_40 = sum(l.amount_ht for l in self.encaissement_ids
                              .filtered(lambda r: r.to_be_delayed is False and r.tax_id.code == 40))  # Code 40
            tva_code_50 = sum(l.amount_ht for l in self.encaissement_ids
                              .filtered(lambda r: r.to_be_delayed is False and r.tax_id.code == 50))  # Code 50
            sum_20_50 = sum([tva_code_20, tva_code_30, tva_code_40, tva_code_50])  # 20 -> 50
            sum_operations = sum_sale_ht - sum_20_50  # Code 060
            sum_sale_tax = sum(
                t.amount_tva for t in self.encaissement_ids.filtered(lambda r: r.to_be_delayed is False))  # Code 130
            sum_purchase_tax = sum(l.amount_tva * (l.tax_id.prorata / 100) for l in self.line_ids
                                   .filtered(lambda r: r.to_be_delayed is False))  # Code 182
            tva_deductible = (sum_purchase_tax + sum(o.value for o in self.manual_field_ids
                                                     .filtered(
                lambda r: r.tva_report_id.code in [170, 180, 181, 185, 186, 187]))) - sum(
                o.value for o in self.manual_field_ids
                    .filtered(lambda r: r.tva_report_id.code in [186, 187]))  # Code 190

            sum_etranger = sum(o.amount_tva for o in self.line_ids
                               .filtered(lambda r: r.tax_id.code in [145]))  # 129

            value = 0
            if line.code == 10:
                value = sum_sale_ht
            # TODO : Sum 20 30 40 50
            elif line.code == 20:
                value = tva_code_20
            elif line.code == 30:
                value = tva_code_30
            elif line.code == 40:
                value = tva_code_40
            elif line.code == 50:
                value = tva_code_50
            elif line.code == 60:
                value = sum_operations
            elif line.code == 130:
                value = sum_sale_tax
            elif line.code == 182:
                value = sum_purchase_tax
            elif line.code == 190:
                value = tva_deductible
            elif line.code == 204:
                value = sum_etranger
            tva_due = sum_sale_tax - sum_etranger - tva_deductible  # 130 - 190 - 129
            
            if line.code == 201 and tva_due < 0  :
                value = abs(tva_due)
            elif line.code == 200  and tva_due > 0 :
                value = tva_due


            if sum_etranger > 0:
                if line.code == 129:
                    value = sum_etranger
                if tva_due > 0:
                    if line.code == 200:
                        value = sum_sale_tax - sum_etranger

                # if tva_due < 0:
                #     if line.code == 130:
                #         value = sum_sale_tax - sum_etranger
                # if line.code == 204:
                #     value = sum_etranger
                # elif line.code == 201:
                #     value = tva_deductible - sum_sale_tax + sum_etranger

            auto_obj = {
                'tva_dec_id': self.id,
                'tva_report_id': line.id,
                'value': value
            }
            self.auto_field_ids.create(auto_obj)

    '''
        Champs à manuels à remplir
    '''

    def get_manual_computed_lines(self):
        self.manual_field_ids.unlink()
        manuel_lines = self.env['account.tax.repport'].search([('type_calcul', '=', 'manuel')])
        dec_mois_prec = self.search([('mois', '=', int(self.period) - 1)])

        for line in manuel_lines:
            value = 0
            if line.code == 170:
                if dec_mois_prec:
                    value = abs(dec_mois_prec.credit_tva)
                    line.value = value
            # manuel_obj = {
            #     'tva_dec_id': self.id,
            #     'tva_report_id': line.id,
            #     'value': value
            # }
            # self.manual_field_ids.create(manuel_obj)

    def get_external_supplier_payments(self):
        external_supplier_obj = self.env['account.external.supplier'].search([('date', '>=', self.period_id.date_start),
                                                                              ('date', '<=', self.period_id.date_end)])
        tva_declaration_line_obj = self.env['tva.declaration.line']
        purchase_tax_id = self.env['account.tax'].search(
            [('code', '=', 145)])  # TODO: replace that with the xml ID of the 129 tax

        seq = 0
        for payment in external_supplier_obj.payment_line_ids:
            seq = seq + 1
            move_id = payment.external_supplier_pay_id.move_line_ids[0].move_id
            pay_method = self.get_payment(move_id)
            invoice_number = move_id.ref
            taux_tva = 0.22221666
            data = {
                'sequence': seq,
                'tva_declaration_id': self.id,
                'move_id': move_id.id,
                'invoice_number': invoice_number,
                'partner_id': move_id.partner_id.id,
                'id_fisc': move_id.partner_id.id_fisc,
                'partner_name': move_id.partner_id.name,
                'ice': move_id.partner_id.ice,
                'description': move_id.invoice_line_ids[0].name,
                # 'amount_ht': tax_dict[tax_line['id']][0],
                'amount_ht': payment.amount,
                'amount_tva': payment.amount * taux_tva,
                'amount_ttc': payment.amount + payment.amount * taux_tva,
                'invoice_date': move_id.date,
                'paiement_date': payment.payment_date,
                'paiement_type': pay_method,
                'tax_rate': 20,
                # 'invoice_tax_id': tax_line['id'],
                'tax_id': purchase_tax_id.id,
                'type': False
            }
            tva_declaration_line_obj.create(data)

    # Generate Data Lines
    def generate_data(self):
        tva_declaration_line_obj = self.env['tva.declaration.line']
        for record in self:
            record.line_ids.filtered(lambda r: r.delayed is False).unlink()
            record.encaissement_ids.filtered(lambda r: r.delayed is False).unlink()
            lines_delayed = tva_declaration_line_obj.search([('to_be_delayed', '=', 'True'),
                                                             ('tva_declaration_id', '!=', record.id)])
            lines_delayed.write({
                'tva_declaration_id': record.id,
                'to_be_delayed': False,
                'delayed': True,
            })
            move_ids = self.get_payment_moves()  # récupération des pièces

            seq_enc = 0
            seq_dec = 0

            for move in move_ids:  # Pièces
                if move.journal_id.default_account_id.user_type_id.type in ('liquidity', 'other'):
                    move_pay_ids = self.generate_liquidity_moves(move)  # récupérer les écritures contre-parties
                    pay_method = self.get_payment(move)

                    for move_pay_line in move_pay_ids:  # Paiement
                        move_pay = move_pay_line[0]
                        move_amount_total = move_pay.move_id.amount_total  # total facture
                        if move_pay.move_id and move_pay.move_id.invoice_line_ids:  # Si le paiment appartient à une/des factures
                            tax_dict = {}

                            total_amount = move_pay_line[1]  # amount payed
                            taux = total_amount / move_amount_total  # taux
                            taxes = []

                            for line in move_pay.move_id.invoice_line_ids:  # lignes de factures
                                price_reduce = line.price_unit * (1.0 - line.discount / 100.0)
                                for tax in line.tax_ids:
                                    taxes.append(tax.compute_all(price_reduce, quantity=line.quantity,
                                                                 product=line.product_id)['taxes'][0])

                            for tax_line in taxes:
                                if total_amount == move_pay.move_id.amount_total:  # totalement rapproché
                                    amount_tax = tax_line['amount']
                                    amount_ht = tax_line['base']
                                    amount_ttc = amount_ht + amount_tax
                                    tax_dict[tax_line['id']] = [amount_ht, amount_tax, amount_ttc]
                                else:  # rapprochement partiel
                                    amount = self.env['account.tax'].search([('id', '=', tax_line['id'])]).amount
                                    amount_tax = taux * tax_line['amount']  # TVA
                                    amount_ht = (amount_tax / amount) * 100
                                    amount_ttc = amount_ht + amount_tax
                                    tax_dict[tax_line['id']] = [amount_ht, amount_tax, amount_ttc]

                            if move_pay.move_id.move_type in ('out_refund', 'in_refund'):  # avoirs
                                tax_dict[tax_line['id']] = [-amount_ht, -amount_tax, -amount_ttc]
                            type = False

                            if move_pay.move_id.move_type in ('out_invoice', 'out_refund'):
                                type = True
                                seq_enc += 1
                            else:
                                seq_dec += 1
                            data = {}
                            for tax_line in taxes:
                                amount = self.env['account.tax'].search([('id', '=', tax_line['id'])]).amount
                                invoice_number = move_pay.move_id.ref
                                move_pay.move_id.declared = True
                                data = {
                                    'sequence': seq_enc,
                                    'tva_declaration_id': record.id,
                                    'move_id': move_pay.move_id.id,
                                    'invoice_number': invoice_number,
                                    'partner_id': move_pay.move_id.partner_id.id,
                                    'id_fisc': move_pay.move_id.partner_id.id_fisc,
                                    'partner_name': move_pay.move_id.partner_id.name,
                                    'ice': move_pay.move_id.partner_id.ice,
                                    'description': move_pay.move_id.invoice_line_ids[0].name,
                                    'amount_ht': tax_dict[tax_line['id']][0],
                                    'amount_tva': tax_dict[tax_line['id']][1],
                                    'amount_ttc': tax_dict[tax_line['id']][2],
                                    'invoice_date': move_pay.move_id.date,
                                    'paiement_date': move.date,
                                    'paiement_type': pay_method,
                                    'tax_rate': amount,
                                    # 'invoice_tax_id': tax_line['id'],
                                    'tax_id': tax_line['id'],
                                    'type': type
                                }
                                if move_pay.move_id.move_type in ('out_invoice', 'out_refund'):
                                    data['sequence'] = seq_enc
                                else:
                                    data['sequence'] = seq_dec
                            tva_declaration_line_obj.create(data)
        self.get_external_supplier_payments()
        self.get_tva_credit()
        self.get_manual_computed_lines()


class TvaDeclarationLine(models.Model):
    _name = "tva.declaration.line"
    _description = "Ligne Declaration Tva"
    _order = 'sequence, id'

    tva_declaration_id = fields.Many2one('tva.declaration', u'Déclaration TVA', ondelete='cascade')
    sequence = fields.Integer(string=u'Séquence')
    move_id = fields.Many2one('account.move', 'Facture', readonly=False)
    invoice_number = fields.Char(u'N°Facture')
    description = fields.Char('Description')
    amount_ht = fields.Float('Montant HT')
    amount_tva = fields.Float('Montant TVA')
    amount_ttc = fields.Float('Montant TTC')
    invoice_date = fields.Date('Date facture')
    paiement_date = fields.Date('Date de paiement')
    paiement_type = fields.Many2one(comodel_name='payement.method',
                                    string=u'Méthode de paiement', readonly=False)
    tax_rate = fields.Float('Taux de TVA', digits=(16, 2))
    # invoice_tax_id = fields.Many2one('account.invoice.tax', 'Taxe facture')
    tax_id = fields.Many2one('account.tax', 'TVA')
    type = fields.Boolean('Encaissement?')
    partner_id = fields.Many2one('res.partner', 'Partenaire')
    id_fisc = fields.Char('IF')
    partner_name = fields.Char('Nom FR')
    ice = fields.Char('ICE')
    to_be_delayed = fields.Boolean('To be delayed?')
    delayed = fields.Boolean('Delayed ?')
    state = fields.Selection(related='tva_declaration_id.state')
    


class TvaLine(models.Model):
    _name = "tva.line"
    _description = "Lignes Tva"

    tva_declaration_id = fields.Many2one('tva.declaration', u'Déclaration TVA', ondelete='cascade')
    tax_id = fields.Many2one('account.tax', 'TVA')
    amount_ht = fields.Float('Base TVA')
    amount = fields.Float('Montant TVA')
    type = fields.Selection([
        ('r', 'Fournisseur'),
        ('p', 'Client')])
    state = fields.Selection(related='tva_declaration_id.state')


class ComputeManualTvaLine(models.Model):
    _name = "tva.compute.manual_line"

    tva_report_id = fields.Many2one('account.tax.repport', 'Ligne TVA')
    value = fields.Float('Montant')
    tva_dec_id = fields.Many2one('tva.declaration')
    state = fields.Selection(related='tva_dec_id.state')


class ComputeAutoTvaLine(models.Model):
    _name = "tva.compute.auto_line"

    tva_report_id = fields.Many2one('account.tax.repport', 'Ligne TVA')
    value = fields.Float('Montant')
    tva_dec_id = fields.Many2one('tva.declaration')
    state = fields.Selection(related='tva_dec_id.state')


# class AccountMoveLine(models.Model):
#     _inherit = 'account.move.line'
#
#     def reconcile(self, writeoff_acc_id=False, writeoff_journal_id=False):
#         fourni_effet_journal = self.env.ref('account_tres_customer.account_journal_data_effar')
#         client_effet_journal = self.env.ref('account_tres_customer.account_journal_data_effap')
#         cheque_journal = self.env.ref('.account_journal_data_chp')
#         print('self.move_id', self.move_id)
#         move_id = self.move_id.filtered(lambda line: line.journal_id in (fourni_effet_journal, client_effet_journal, cheque_journal))
#         print('move_id', move_id)
#         if move_id:
#             invoice_move_id = self.move_id.filtered(lambda line: line != move_id)
#             move_id.write({
#                 'date': invoice_move_id.invoice_date
#             })
#         res = super(AccountMoveLine, self).reconcile(writeoff_acc_id=writeoff_acc_id,
#                                                      writeoff_journal_id=writeoff_journal_id)
#         return res
